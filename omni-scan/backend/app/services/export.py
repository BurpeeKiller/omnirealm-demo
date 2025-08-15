"""Service d'export multi-formats pour OmniScan"""

import json
import io
from datetime import datetime
from typing import Dict, Any, List
from reportlab.lib.pagesizes import A4
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER, TA_JUSTIFY
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment


class ExportService:
    """Service pour exporter les documents OCR dans différents formats"""
    
    @staticmethod
    async def export_to_json(document_data: Dict[str, Any]) -> bytes:
        """Exporter en format JSON"""
        export_data = {
            "document_id": document_data.get("id"),
            "filename": document_data.get("filename"),
            "processed_date": document_data.get("created_at", datetime.now().isoformat()),
            "ocr_text": document_data.get("ocr_text", ""),
            "metadata": {
                "file_type": document_data.get("file_type"),
                "file_size": document_data.get("file_size"),
                "pages": document_data.get("pages", 1),
                "language": document_data.get("language", "fra"),
                "confidence_score": document_data.get("confidence_score")
            },
            "ai_analysis": document_data.get("ai_analysis", {})
        }
        
        return json.dumps(export_data, ensure_ascii=False, indent=2).encode('utf-8')
    
    @staticmethod
    async def export_to_excel(document_data: Dict[str, Any]) -> bytes:
        """Exporter en format Excel avec mise en forme"""
        wb = Workbook()
        ws = wb.active
        ws.title = "OCR Results"
        
        # Styles
        subheader_font = Font(bold=True, size=12)
        
        # En-tête principal
        ws['A1'] = 'OmniScan OCR Export'
        ws['A1'].font = Font(bold=True, size=16)
        ws.merge_cells('A1:D1')
        ws['A1'].alignment = Alignment(horizontal='center')
        
        # Informations du document
        info_data = [
            ['', '', '', ''],
            ['Document Information', '', '', ''],
            ['Filename:', document_data.get('filename', 'N/A'), 'Document ID:', document_data.get('id', 'N/A')],
            ['File Type:', document_data.get('file_type', 'N/A'), 'File Size:', f"{document_data.get('file_size', 0):,} bytes"],
            ['Processed Date:', document_data.get('created_at', datetime.now().strftime('%Y-%m-%d %H:%M')), 'Pages:', document_data.get('pages', 1)],
            ['', '', '', '']
        ]
        
        for row in info_data:
            ws.append(row)
        
        # Mise en forme de la section info
        ws['A3'].font = subheader_font
        ws.merge_cells('A3:D3')
        
        # Texte OCR
        ws.append(['OCR Extracted Text', '', '', ''])
        ocr_row = ws.max_row
        ws[f'A{ocr_row}'].font = subheader_font
        ws.merge_cells(f'A{ocr_row}:D{ocr_row}')
        
        # Diviser le texte en lignes pour une meilleure lisibilité
        ocr_text = document_data.get('ocr_text', '')
        text_lines = ocr_text.split('\n')
        
        for line in text_lines[:100]:  # Limiter à 100 lignes pour Excel
            if line.strip():
                ws.append([line, '', '', ''])
                ws.merge_cells(f'A{ws.max_row}:D{ws.max_row}')
        
        if len(text_lines) > 100:
            ws.append(['... (texte tronqué, voir export PDF pour le texte complet)', '', '', ''])
            ws.merge_cells(f'A{ws.max_row}:D{ws.max_row}')
        
        # Analyse IA si disponible
        if document_data.get('ai_analysis'):
            ws.append(['', '', '', ''])
            ws.append(['AI Analysis Results', '', '', ''])
            ai_row = ws.max_row
            ws[f'A{ai_row}'].font = subheader_font
            ws.merge_cells(f'A{ai_row}:D{ai_row}')
            
            ai_data = document_data.get('ai_analysis', {})
            if isinstance(ai_data, dict):
                for key, value in ai_data.items():
                    ws.append([key.replace('_', ' ').title() + ':', str(value), '', ''])
                    ws.merge_cells(f'B{ws.max_row}:D{ws.max_row}')
        
        # Ajuster les largeurs de colonnes
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 40
        ws.column_dimensions['C'].width = 20
        ws.column_dimensions['D'].width = 20
        
        # Sauvegarder en mémoire
        excel_buffer = io.BytesIO()
        wb.save(excel_buffer)
        excel_buffer.seek(0)
        
        return excel_buffer.getvalue()
    
    @staticmethod
    async def export_to_pdf(document_data: Dict[str, Any]) -> bytes:
        """Exporter en format PDF avec mise en forme professionnelle"""
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4)
        story = []
        styles = getSampleStyleSheet()
        
        # Styles personnalisés
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Title'],
            fontSize=24,
            textColor=colors.HexColor('#1a73e8'),
            spaceAfter=30,
            alignment=TA_CENTER
        )
        
        heading_style = ParagraphStyle(
            'CustomHeading',
            parent=styles['Heading2'],
            fontSize=14,
            textColor=colors.HexColor('#1a73e8'),
            spaceAfter=12,
            spaceBefore=20
        )
        
        # Titre
        story.append(Paragraph("OmniScan OCR Export Report", title_style))
        story.append(Spacer(1, 0.5*inch))
        
        # Informations du document
        story.append(Paragraph("Document Information", heading_style))
        
        doc_info = [
            ['Filename:', document_data.get('filename', 'N/A')],
            ['Document ID:', document_data.get('id', 'N/A')],
            ['File Type:', document_data.get('file_type', 'N/A')],
            ['File Size:', f"{document_data.get('file_size', 0):,} bytes"],
            ['Processed Date:', document_data.get('created_at', datetime.now().strftime('%Y-%m-%d %H:%M'))],
            ['Pages:', str(document_data.get('pages', 1))],
            ['Language:', document_data.get('language', 'French')]
        ]
        
        doc_table = Table(doc_info, colWidths=[2*inch, 4*inch])
        doc_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (0, -1), colors.HexColor('#f0f0f0')),
            ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('PADDING', (0, 0), (-1, -1), 8),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey)
        ]))
        story.append(doc_table)
        story.append(Spacer(1, 0.5*inch))
        
        # Texte OCR
        story.append(Paragraph("Extracted Text", heading_style))
        
        ocr_text = document_data.get('ocr_text', '')
        # Diviser le texte en paragraphes pour une meilleure mise en forme
        paragraphs = ocr_text.split('\n\n')
        
        text_style = ParagraphStyle(
            'TextStyle',
            parent=styles['BodyText'],
            fontSize=11,
            alignment=TA_JUSTIFY,
            spaceAfter=12
        )
        
        for para in paragraphs:
            if para.strip():
                # Échapper les caractères spéciaux pour ReportLab
                safe_para = para.replace('&', '&amp;').replace('<', '&lt;').replace('>', '&gt;')
                story.append(Paragraph(safe_para, text_style))
        
        # Analyse IA si disponible
        if document_data.get('ai_analysis'):
            story.append(Spacer(1, 0.5*inch))
            story.append(Paragraph("AI Analysis Results", heading_style))
            
            ai_data = document_data.get('ai_analysis', {})
            if isinstance(ai_data, dict):
                ai_info = []
                for key, value in ai_data.items():
                    formatted_key = key.replace('_', ' ').title()
                    ai_info.append([formatted_key + ':', str(value)])
                
                if ai_info:
                    ai_table = Table(ai_info, colWidths=[2*inch, 4*inch])
                    ai_table.setStyle(TableStyle([
                        ('BACKGROUND', (0, 0), (0, -1), colors.HexColor('#e8f4f8')),
                        ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
                        ('FONTSIZE', (0, 0), (-1, -1), 10),
                        ('PADDING', (0, 0), (-1, -1), 8),
                        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey)
                    ]))
                    story.append(ai_table)
        
        # Footer
        story.append(Spacer(1, inch))
        footer_style = ParagraphStyle(
            'Footer',
            parent=styles['BodyText'],
            fontSize=9,
            textColor=colors.grey,
            alignment=TA_CENTER
        )
        story.append(Paragraph(f"Generated by OmniScan Pro • {datetime.now().strftime('%Y-%m-%d %H:%M')}", footer_style))
        
        # Construire le PDF
        doc.build(story)
        buffer.seek(0)
        
        return buffer.getvalue()
    
    @staticmethod
    async def export_batch(documents: List[Dict[str, Any]], format: str = "json") -> bytes:
        """Exporter plusieurs documents en un seul fichier"""
        if format == "json":
            export_data = {
                "export_date": datetime.now().isoformat(),
                "total_documents": len(documents),
                "documents": []
            }
            
            for doc in documents:
                doc_data = {
                    "document_id": doc.get("id"),
                    "filename": doc.get("filename"),
                    "processed_date": doc.get("created_at"),
                    "ocr_text": doc.get("ocr_text", ""),
                    "metadata": {
                        "file_type": doc.get("file_type"),
                        "file_size": doc.get("file_size"),
                        "pages": doc.get("pages", 1)
                    }
                }
                export_data["documents"].append(doc_data)
            
            return json.dumps(export_data, ensure_ascii=False, indent=2).encode('utf-8')
        
        elif format == "excel":
            wb = Workbook()
            
            # Feuille de résumé
            summary_ws = wb.active
            summary_ws.title = "Summary"
            summary_ws['A1'] = 'OmniScan Batch Export'
            summary_ws['A1'].font = Font(bold=True, size=16)
            summary_ws['A3'] = 'Total Documents:'
            summary_ws['B3'] = len(documents)
            summary_ws['A4'] = 'Export Date:'
            summary_ws['B4'] = datetime.now().strftime('%Y-%m-%d %H:%M')
            
            # Une feuille par document
            for idx, doc in enumerate(documents[:10]):  # Limiter à 10 documents
                ws = wb.create_sheet(title=f"Doc_{idx+1}")
                ws['A1'] = doc.get('filename', f'Document {idx+1}')
                ws['A1'].font = Font(bold=True, size=14)
                ws.merge_cells('A1:D1')
                
                # Texte OCR
                ws['A3'] = 'OCR Text:'
                ws['A3'].font = Font(bold=True)
                
                text_lines = doc.get('ocr_text', '').split('\n')
                row = 4
                for line in text_lines[:50]:  # Limiter les lignes
                    if line.strip():
                        ws[f'A{row}'] = line
                        ws.merge_cells(f'A{row}:D{row}')
                        row += 1
            
            excel_buffer = io.BytesIO()
            wb.save(excel_buffer)
            excel_buffer.seek(0)
            
            return excel_buffer.getvalue()
        
        else:
            raise ValueError(f"Format non supporté: {format}")